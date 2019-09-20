from django.shortcuts import render
from django.template import loader, Context
from django.db.models import Q, Max
from django.utils import timezone
import openpyxl
import numpy as np
import os, io, pytz
from datetime import datetime, timedelta
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt 
from django.http import HttpResponse
from django.apps import apps

from .models import TempHumidData, Room, Interval
from .forms import RoomForm, IntervalForm

# Create your views here.

def simulate_temperature(room_query):
    """
    Arguments:
    room_query: query of django model Room
    """
    cur_temp_dict = {}
    cur_humid_dict = {}
    for i, room in enumerate(room_query):
        cur_temp_dict[room] = np.random.randn()*4+(i/2+6)*10
        cur_humid_dict[room] = np.random.randn()*4+(i/2+2)*10
    return [cur_temp_dict, cur_humid_dict]
    
def index(request):
    # Obtain user info 
    User = apps.get_model('home', 'User')
    user_dt = User.objects.get(username=request.user.username)
    datas = TempHumidData.objects.filter(parent_user = user_dt)
    rooms = Room.objects.filter(parent_user = user_dt)
    interval_dt = Interval.objects.first()
    
    if interval_dt is None:
        interval_dt = Interval(interval=10, parent_user=user_dt)
    
    cur_datetime = timezone.now()
    cur_datetime_str = cur_datetime.strftime("%c")
    last_datetime = datas.aggregate(Max('datetime'))['datetime__max']
    # delete data more than one day old
    datas.filter(datetime__lte = cur_datetime - timedelta(0, 24*60*60)).delete()
            
    
    
    def add_newdata(target_rooms):
        """
        Add new data to the query
        Argument:
        target_rooms: query of rooms
        return:
        the new query of TempHumidData
        """
        new_simulation = simulate_temperature(target_rooms)
        for room in target_rooms:
            trunc = cur_datetime.replace(minute=cur_datetime.minute-cur_datetime.minute%interval, second=0, microsecond=0)
            new_data = TempHumidData(datetime=trunc,
                                     temperature=new_simulation[0][room], humidity=new_simulation[1][room], parent_user=user_dt, parent_room=room)
            new_data.save()
        return TempHumidData.objects.filter(parent_user = user_dt)
    
    interval = interval_dt.interval
    
    # Add a new room
    if request.method == 'POST':
        new_room = Room(parent_user=user_dt)
        new_int = Interval(parent_user = user_dt)
        roomform = RoomForm(request.POST, instance=new_room)
        intform = IntervalForm(request.POST, instance=new_int)
        if roomform.is_valid():
            roomform.save()
            rooms = Room.objects.filter(parent_user = user_dt)
            datas = add_newdata([new_room])
        elif intform.is_valid():
            intform.save()
            interval_dt.delete()
            interval_dt = new_int
    roomform = RoomForm()
    intform = IntervalForm()
            
    interval = interval_dt.interval
    
    # Every interval min, obtain new temp/humid
    if not datas or last_datetime is None:
        datas = add_newdata(rooms)
    elif (cur_datetime - last_datetime).seconds > 60*interval:
        datas = add_newdata(rooms)
    
    # delete duplicates
    for room in rooms.distinct():
        rooms.filter(pk__in=rooms.filter(name=room.name).values_list('id', flat=True)[1:]).delete()
        
    
    time_dict = {}
    time_str_dict = {}
    temp_dict = {}
    humid_dict = {}
    for room in rooms:
        target_datas = datas.filter(parent_room = room)
        time_dict[room] = [dat.datetime for dat in target_datas]
        time_str_dict[room] = [dat.datetime.strftime("%H:%M") for dat in target_datas]
        temp_dict[room] = [dat.temperature for dat in target_datas]
        humid_dict[room] = [dat.humidity for dat in target_datas]
        
    # complete list of time supposed to be displayed
    ref_time_list = [(cur_datetime.replace(hour=0, minute=0, second=0, microsecond=0)+timedelta(0, i*60*interval)) for i in range((60//interval)*24+1)]
    ref_time_list_str = [t.strftime("%H:%M") for t in ref_time_list]
    # Fill None to the missing data
    for room in rooms:
        for t in ref_time_list:
            if t not in time_dict[room]:
                time_dict[room].append(t)
                time_dict[room].sort()
                target_time_str = [t.strftime("%H:%M") for t in time_dict[room]]
                time_str_dict[room] = target_time_str
                idx = time_dict[room].index(t)
                temp_dict[room].insert(idx, np.nan)
                humid_dict[room].insert(idx, np.nan)
            else:
                print(t)
                
    """
    # Using excel file is deprecated. 
    for row in range(2,worksheet.max_row):
        if(worksheet.cell(row,1).value is not None):
            row_cells = [worksheet.cell(row,col) for col in range(2, worksheet.max_column+1)]
        row_data = list()
        for col, cell in enumerate(row_cells):
            if col == 0:
                dt = str(cell.value)[-8:]
                datetime_object = datetime.strptime(dt, '%H:%M:%S')
                time_data.append(datetime_object.strftime("%H:%M"))
            else:
                row_data.append(float(cell.value))
        excel_data.append(row_data)

    excel_data = np.array(excel_data).T
    """
    print(len(ref_time_list_str))
    def setPlt(x_list, y_list, label):
        plt.plot(range(len(x_list)), y_list, label=label, marker='o')
    def pltToSvg(buf):
        plt.savefig(buf, format='svg', bbox_inches='tight')
        s = buf.getvalue()
        return s
    
    for room in rooms:
        setPlt(ref_time_list_str, temp_dict[room], room)
    plt.xlabel('Time')
    plt.ylabel('Temperature ($^\circ$F)')
    plt.xticks(range(len(ref_time_list_str)), ref_time_list_str[::10], rotation=45, ha="right")
    plt.xlim([-1, (60//interval)*24+2])
    plt.locator_params(axis='x', nbins=((60//interval)*24+1)/10)
    plt.legend()
    buf1 = io.StringIO()
    svg1 = pltToSvg(buf1)
    plt.close()
    
    for room in rooms:
        setPlt(ref_time_list_str, humid_dict[room], room)
    plt.xlabel('Time')
    plt.ylabel('Humidity (%)')
    plt.xticks(range(len(ref_time_list_str)), ref_time_list_str[::10], rotation=45, ha="right")
    plt.xlim([-1, (60//interval)*24+2])
    plt.locator_params(axis='x', nbins=((60//interval)*24+1)/10)
    plt.legend()
    buf2 = io.StringIO() 
    svg2 = pltToSvg(buf2)
    plt.close()
      
    template = loader.get_template('temperature/index.html')
    response = HttpResponse(template.render({'cur_datetime': cur_datetime_str, 'svg1': svg1, 'svg2': svg2, 'roomform': roomform, 'intform': intform, 'interval': interval}, request))
    buf1.close()
    buf2.close()
    return response